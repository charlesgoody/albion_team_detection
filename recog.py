import os
from paddleocr import PaddleOCR
from scrapper import scrapper
import openpyxl
from openpyxl.styles import Font
import sys
from rapidfuzz import process, fuzz






if __name__ == "__main__":

    message = """
    ┌─str──────────────────────────────────────────────────────────────────────┐
    |                                                                          |
    │ Players.exe created by Char1es. Mail :dakeleecharles7827@gmail.com       │
    |                                                                          |
    └──────────────────────────────────────────────────────────────────────────┘

    """


    done = """
    ┌─str───────────────┐
    │ Task complete.    │
    └───────────────────┘
    """

    prcrossing = """
    ┌─str────────────┐
    │ Processing.    │
    └────────────────┘
    """

    print(message)
    root = os.path.dirname(os.path.abspath(sys.argv[0]))
    png_files = [i for i in os.listdir(root) if i.endswith(".png")]
   
    xlsx_path = "attendance.xlsx"
    workbook = openpyxl.Workbook()  
    column = 1
    ws = workbook.worksheets[0]
    ws.title = "Each_pics"
    w2 = workbook.create_sheet("Total")

    x = input("Your guild name:")
    search_url = scrapper.get_guild_name(x)

    member_list = scrapper.get_all_members(search_url)

    print(member_list)
    

    for index, member in enumerate(member_list):
        w2.cell(row=index+1,column=1).value = member
        w2.cell(row=index+1,column=2).value = 0

    print(png_files)

    print(prcrossing)

    

    ocr = PaddleOCR(lang="en",det_db_thresh=0.3,det_db_box_thresh=0.4,use_dilation=True,show_log=False,drop_score=0.65)


   
    
    for png_file in png_files:
        attendence = list()
        error_list = list()
        path = os.path.join(root, png_file)
        result = ocr.ocr(path,cls=False, inv=True)
        print(f" Now recognize :   {png_file} ")
        for idx in range(len(result)):
            res = result[idx]
            print(res)
            for line in res:
                if line[1][0] in member_list:           
                    attendence.append([line[1][0],line[1][1]])
                else:
                    error_list.append(line[1][0])
                    

        for i in range(len(error_list)):
            item = error_list[i]
            match, score, _ = process.extractOne(item, member_list, scorer=fuzz.ratio)
            if score > 70:
                attendence.append([match,score])
                
        
        
        ws.cell(row=1,column=column).font = Font(color="FF0000")
        ws.cell(row=1,column=column).border = None
        ws.cell(row=1,column=column).value = png_file
        
        for i , value in  enumerate(attendence):
            ws.cell(row=i+3,column=column).value = value[0]
        
        ws.cell(row=len(attendence)+3,column=column).value = len(attendence)
        
        column +=2
           

        print(attendence)

        attendance_set = set()
        for i in attendence:
            attendance_set.add(i[0])
        

        for row in range(1,w2.max_row + 1):
            if w2.cell(row = row,column = 1).value in attendance_set:
                w2.cell(row = row,column = 2).value += 1

        print("Collect data from {}".format(png_file))

    
    workbook.save(xlsx_path)

    print(done)               
                    



